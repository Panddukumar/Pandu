import openpyxl
from openpyxl.styles import PatternFill


def getrowcount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row

def getcolumncount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_column

def readdata(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rownum,colnum).value

def writedata(file,sheetname,rownum,colnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(rownum,colnum).value=data
    workbook.save(file)

def greencolor(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    greenfill=PatternFill(start_color='60b212',
                          end_color='60b212',
                          fill_type='solid')
    sheet.cell(rownum,colnum).fill=greenfill
    workbook.save(file)
    

def redcolor(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    redfill=PatternFill(start_color='ff0000',
                        end_color='ff0000',
                        fill_type='solid')
    sheet.cell(rownum,colnum).fill=redfill
    workbook.save(file)
